# -*- coding: utf-8 -*-
"""
Proyectista de racks — PM La Piedad / Grupo PEME.

Este módulo implementa el agente completo descrito en las instrucciones del
proyectista: prompt de sistema, catálogo, validador de reglas estructurales,
y generadores de PDF de planos, XLSX de despiece/cotización, y render 3D
determinista en HTML/Three.js.

Se mantiene separado de main.py para no inflar más ese archivo; main.py solo
importa y orquesta (llamar al modelo, guardar en Supabase, subir a Storage).

⚠️ IMPORTANTE: `CATALOGO_PM` de este archivo es un PLACEHOLDER con los códigos
que menciona el documento de instrucciones, pero SIN precios reales (no los
tengo). Reemplázalo por tu `catalogo_pm.json` real, o carga el catálogo real
en la tabla `catalogo_pm` de Supabase (ver `consultar_catalogo_pm`, que
prioriza Supabase sobre este placeholder).
"""

import io
import json
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas




# ⚠️ Precios en None = PENDIENTE. Dimensiones puestas a modo de ejemplo
# respetando los valores válidos del validador, pero NO son el catálogo real
# de PM. Súbelo a la tabla `catalogo_pm` de Supabase o reemplaza esta lista.

FRENTES_VALIDOS = {1294, 1594, 1894, 2294, 2504, 2804, 3104}
FONDOS_VALIDOS = {612, 917, 1232}
ALTURAS_CABECERA_CATALOGO = [1226, 1530, 1834, 2240, 2443, 2748, 3001, 3357, 3665, 4025]
PERALTES_PESADA = {100, 125, 150}
PERALTE_LIGERA = 75
CAP_MARCO_PESADA_KG = 4500
CAP_MARCO_LIGERA_KG = 2500
FACTOR_SEGURIDAD_MIN = 1.5


def consultar_catalogo_pm(supabase_client) -> list[dict]:
    """
    Trae el catálogo real de PM desde Supabase (tabla `catalogo_pm`) si existe
    y tiene datos; si no, usa el placeholder de este archivo (sin precios).
    """
    try:
        resultado = supabase_client.table("catalogo_pm").select("*").execute()
        if resultado.data and len(resultado.data) > 0:
            return resultado.data
    except Exception:
        pass
    return CATALOGO_PM_PLACEHOLDER


def construir_system_prompt_pm(catalogo: list[dict], fichas_tecnicas: str = "") -> str:
    """
    Arma el system_prompt completo: instrucciones base + catálogo disponible +
    fichas técnicas (si se proporcionan; Fase 1 pueden ir vacías/embebidas).
    """
    prompt = SYSTEM_PROMPT_PM_BASE
    prompt += f"\n\n## CATÁLOGO DISPONIBLE (`catalogo_pm.json`)\n```json\n{json.dumps(catalogo, indent=2, ensure_ascii=False)}\n```\n"
    if fichas_tecnicas:
        prompt += f"\n## FICHAS TÉCNICAS (knowledge/tecnico/)\n{fichas_tecnicas}\n"
    else:
        prompt += (
            "\n## FICHAS TÉCNICAS (knowledge/tecnico/)\n"
            "(No hay fichas técnicas cargadas todavía. Usa las reglas estructurales "
            "de este prompt como única fuente de verdad mientras tanto.)\n"
        )
    return prompt


# ============================================================================
# 3. EXTRACCIÓN DEL JSON DESDE LA RESPUESTA DE CLAUDE
# ============================================================================

def extraer_json_proyecto(texto_respuesta: str) -> tuple[dict | None, str]:
    """
    Separa la respuesta de Claude en (json_del_proyecto, texto_narrativo).
    El texto narrativo es todo lo anterior al bloque ```json (secciones 1-4:
    Diseño propuesto, Supuestos, Despiece, Cotización) — eso es lo que se le
    manda al cliente por Telegram/dashboard tal cual.
    """
    match = re.search(r"```json\s*(\{.*?\})\s*```", texto_respuesta, re.DOTALL)
    if not match:
        return None, texto_respuesta.strip()

    texto_narrativo = texto_respuesta[: match.start()].strip()
    try:
        datos = json.loads(match.group(1))
    except json.JSONDecodeError:
        datos = None
    return datos, texto_narrativo


# ============================================================================
# 4. VALIDADOR DE REGLAS ESTRUCTURALES
# ============================================================================

def _es_pesada(especificacion: str) -> bool:
    return "pesada" in (especificacion or "").lower()


def _es_ligera(especificacion: str) -> bool:
    return "ligera" in (especificacion or "").lower()



# ============================================================================
# 5. GENERADOR DE XLSX — DESPIECE Y COTIZACIÓN
# ============================================================================

_FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_FILL_HEADER = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
_FONT_BODY = Font(name="Calibri", size=10)
_BORDER_THIN = Border(*(Side(style="thin", color="D1D5DB"),) * 4)


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 3, 10), 60)


def generar_xlsx_despiece(datos: dict) -> bytes:
    """Genera el XLSX de despiece: Pzas | Código | Descripción | Color."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Despiece"

    ws.append([f"Proyecto: {datos.get('proyecto', '')}", "", f"Clave: {datos.get('clave', '')}", ""])
    ws.append([f"Cliente: {datos.get('cliente', '')}", "", f"Fecha: {datos.get('fecha', '')}", ""])
    ws.append([])

    headers = ["Pzas", "Código", "Descripción", "Color", "Obs"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")

    for m in datos.get("materiales", []) or []:
        ws.append([
            m.get("pzas", 0), m.get("codigo", ""), m.get("descripcion", ""),
            m.get("color", ""), m.get("obs", ""),
        ])
        for cell in ws[ws.max_row]:
            cell.font = _FONT_BODY
            cell.border = _BORDER_THIN

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_xlsx_cotizacion(datos: dict) -> bytes:
    """Genera el XLSX de cotización: Código | Descripción | Cant | P.Unit | Importe + subtotal."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    ws.append([f"Proyecto: {datos.get('proyecto', '')}", "", f"Clave: {datos.get('clave', '')}"])
    ws.append([f"Cliente: {datos.get('cliente', '')}", "", f"Fecha: {datos.get('fecha', '')}"])
    ws.append([])

    headers = ["Código", "Descripción", "Cant", "P.Unit (MXN)", "Importe (MXN)"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")

    subtotal = 0.0
    for m in datos.get("materiales", []) or []:
        pzas = m.get("pzas", 0) or 0
        precio = m.get("precio") or 0
        importe = pzas * precio
        subtotal += importe
        ws.append([m.get("codigo", ""), m.get("descripcion", ""), pzas, precio, importe])
        row = ws[ws.max_row]
        row[3].number_format = '"$"#,##0.00'
        row[4].number_format = '"$"#,##0.00'
        for cell in row:
            cell.font = _FONT_BODY
            cell.border = _BORDER_THIN

    ws.append([])
    ws.append(["", "", "", "Subtotal (MXN, s/IVA):", subtotal])
    subtotal_row = ws[ws.max_row]
    subtotal_row[3].font = Font(bold=True)
    subtotal_row[4].font = Font(bold=True)
    subtotal_row[4].number_format = '"$"#,##0.00'
    ws.append(["", "", "", "IVA, flete e instalación se cotizan aparte.", ""])
    ws[ws.max_row][3].font = Font(italic=True, size=9, color="64748B")

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================================
# 6. GENERADOR DE PDF DE PLANOS (vista superior determinista)
# ============================================================================

def generar_pdf_planos(datos: dict) -> bytes:
    """
    Dibuja un plano en vista superior (top-down) desde `layout`: corridas,
    bays, pasillo central, con acotación y cuadro de datos del proyecto.
    """
    layout = datos.get("layout", {}) or {}
    n_bays = layout.get("modulos_x", 1) or 1
    n_corridas = layout.get("modulos_y", 1) or 1
    frente_mm = layout.get("frente_mm", 2000) or 2000
    fondo_mm = layout.get("fondo_mm", 1000) or 1000
    pasillo_mm = layout.get("pasillo_mm", 3000) or 3000

    buf = io.BytesIO()
    page_w, page_h = landscape(A4)
    c = pdf_canvas.Canvas(buf, pagesize=landscape(A4))

    margin = 15 * mm
    dibujo_w = page_w - 2 * margin
    dibujo_h = page_h - 2 * margin - 30 * mm  # deja espacio al cuadro de datos

    ancho_total_mm = n_bays * frente_mm
    fondo_total_mm = n_corridas * fondo_mm + max(n_corridas - 1, 0) * pasillo_mm
    escala = min(dibujo_w / ancho_total_mm, dibujo_h / fondo_total_mm) * 0.92

    origen_x = margin + (dibujo_w - ancho_total_mm * escala) / 2
    origen_y = margin + 30 * mm

    # --- Título ---
    c.setFont("Helvetica-Bold", 14)
    c.drawString(margin, page_h - margin, f"{datos.get('proyecto', 'Proyecto')} — Plano en planta")
    c.setFont("Helvetica", 9)
    c.drawString(margin, page_h - margin - 14, f"Clave: {datos.get('clave', '—')}   Cliente: {datos.get('cliente', '—')}   Fecha: {datos.get('fecha', '—')}   Rev: {datos.get('revision', '—')}")

    # --- Corridas y bays ---
    c.setLineWidth(1)
    y_cursor = origen_y
    for corrida in range(n_corridas):
        x_cursor = origen_x
        for bay in range(n_bays):
            c.setFillColorRGB(0.93, 0.93, 0.97)
            c.rect(x_cursor, y_cursor, frente_mm * escala, fondo_mm * escala, fill=1, stroke=1)
            c.setFillColorRGB(0, 0, 0)
            x_cursor += frente_mm * escala
        # Etiqueta de corrida
        c.setFont("Helvetica-Bold", 8)
        c.drawString(origen_x - 12 * mm, y_cursor + (fondo_mm * escala) / 2 - 3, f"C{corrida + 1}")
        y_cursor += fondo_mm * escala
        if corrida < n_corridas - 1:
            # Pasillo entre corridas
            c.setDash(3, 3)
            c.setStrokeColorRGB(0.6, 0.6, 0.6)
            c.rect(origen_x, y_cursor, ancho_total_mm * escala, pasillo_mm * escala, fill=0, stroke=1)
            c.setStrokeColorRGB(0, 0, 0)
            c.setDash()
            c.setFont("Helvetica-Oblique", 7)
            c.drawCentredString(origen_x + (ancho_total_mm * escala) / 2, y_cursor + (pasillo_mm * escala) / 2 - 3,
                                 f"Pasillo {pasillo_mm}mm")
            y_cursor += pasillo_mm * escala

    # --- Acotación general ---
    c.setFont("Helvetica", 7)
    c.drawCentredString(origen_x + (ancho_total_mm * escala) / 2, origen_y - 10,
                         f"Frente total: {ancho_total_mm}mm ({n_bays} módulos x {frente_mm}mm)")
    c.saveState()
    c.translate(origen_x - 22, origen_y + (fondo_total_mm * escala) / 2)
    c.rotate(90)
    c.drawCentredString(0, 0, f"Fondo total: {fondo_total_mm}mm")
    c.restoreState()

    # --- Cuadro de datos ---
    cuadro_y = margin
    c.setFont("Helvetica-Bold", 8)
    c.drawString(margin, cuadro_y + 12, "Elaboró:")
    c.drawString(margin + 110, cuadro_y + 12, "Revisó:")
    c.drawString(margin + 220, cuadro_y + 12, "Aprobó:")
    c.setFont("Helvetica", 8)
    c.drawString(margin, cuadro_y, str(datos.get("elaboro", "—")))
    c.drawString(margin + 110, cuadro_y, str(datos.get("reviso", "—")))
    c.drawString(margin + 220, cuadro_y, str(datos.get("aprobo", "—")))
    c.drawString(margin + 340, cuadro_y + 12, f"Especificación: {datos.get('especificacion', '—')}")
    c.drawString(margin + 340, cuadro_y, f"Material: {datos.get('material', '—')}  {datos.get('calibre', '')}")

    c.showPage()
    c.save()
    return buf.getvalue()


# ============================================================================
# 7. RENDER 3D DETERMINISTA (HTML + Three.js)
# ============================================================================

def generar_render_3d_html(datos: dict) -> str:
    """
    Construye un render 3D determinista en HTML/Three.js a partir de `layout`.
    No es CAD de fabricación pixel-perfect, pero refleja fielmente: número de
    marcos/corridas/bays, alturas de niveles, postes 73mm (pesada) / 38mm
    (ligera), largueros por nivel, cargadores extra si frente >= 2804mm, y
    cruces (x-bracing) en zigzag entre postes frontal/trasero de cada marco.
    """
    layout = datos.get("layout", {}) or {}
    especificacion = datos.get("especificacion", "") or ""
    pesada = _es_pesada(especificacion)

    n_bays = layout.get("modulos_x", 1) or 1
    n_corridas = layout.get("modulos_y", 1) or 1
    frente_mm = layout.get("frente_mm", 2000) or 2000
    fondo_mm = layout.get("fondo_mm", 1000) or 1000
    pasillo_mm = layout.get("pasillo_mm", 3000) or 3000
    niveles = layout.get("niveles") or [0]
    altura_total = layout.get("altura_total_mm") or max(niveles)

    radio_poste = 0.073 if pesada else 0.038
    color_poste = "0xd97706" if pesada else "0x2563eb"
    dos_cargadores = frente_mm >= 2804

    # Convertimos a metros para Three.js
    m_frente = frente_mm / 1000
    m_fondo = fondo_mm / 1000
    m_pasillo = pasillo_mm / 1000
    m_niveles = [n / 1000 for n in niveles]
    m_altura = altura_total / 1000

    geometria_json = json.dumps({
        "n_bays": n_bays, "n_corridas": n_corridas,
        "frente": m_frente, "fondo": m_fondo, "pasillo": m_pasillo,
        "niveles": m_niveles, "altura": m_altura,
        "radio_poste": radio_poste, "color_poste": color_poste,
        "dos_cargadores": dos_cargadores, "pesada": pesada,
    })

    proyecto_nombre = (datos.get("proyecto") or "Proyecto").replace("<", "").replace(">", "")

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Render 3D — {proyecto_nombre}</title>
<style>
  html, body {{ margin:0; height:100%; background:#0f172a; overflow:hidden; font-family: sans-serif; }}
  #info {{ position:absolute; top:10px; left:10px; color:#e2e8f0; font-size:12px; z-index:10; background:rgba(15,23,42,.7); padding:8px 12px; border-radius:8px; }}
</style>
</head>
<body>
<div id="info">{proyecto_nombre} — render determinista desde layout</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/three.js/r128/three.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/three@0.128.0/examples/js/controls/OrbitControls.js"></script>
<script>
const G = {geometria_json};

const scene = new THREE.Scene();
scene.background = new THREE.Color(0x0f172a);

const camera = new THREE.PerspectiveCamera(50, window.innerWidth/window.innerHeight, 0.1, 1000);
const anchoTotal = G.n_bays * G.frente;
const fondoTotal = G.n_corridas * G.fondo + Math.max(G.n_corridas-1,0) * G.pasillo;
camera.position.set(anchoTotal*0.9, G.altura*1.6, fondoTotal*1.6 + anchoTotal*0.4);

const renderer = new THREE.WebGLRenderer({{antialias:true}});
renderer.setSize(window.innerWidth, window.innerHeight);
document.body.appendChild(renderer.domElement);

const controls = new THREE.OrbitControls(camera, renderer.domElement);
controls.target.set(anchoTotal/2, G.altura/2, fondoTotal/2);
controls.update();

scene.add(new THREE.AmbientLight(0xffffff, 0.6));
const dir = new THREE.DirectionalLight(0xffffff, 0.8);
dir.position.set(5, 10, 5);
scene.add(dir);

const matPoste = new THREE.MeshStandardMaterial({{color: parseInt(G.color_poste)}});
const matViga = new THREE.MeshStandardMaterial({{color: 0x94a3b8}});
const matCruz = new THREE.MeshStandardMaterial({{color: 0x475569}});
const matPiso = new THREE.MeshStandardMaterial({{color: 0x1e293b}});

scene.add(new THREE.Mesh(new THREE.PlaneGeometry(anchoTotal*1.4, fondoTotal*1.4), matPiso).rotateX(-Math.PI/2).translateZ(0));

function poste(x, y, z, alto) {{
  const geo = new THREE.CylinderGeometry(G.radio_poste, G.radio_poste, alto, 12);
  const mesh = new THREE.Mesh(geo, matPoste);
  mesh.position.set(x, y + alto/2, z);
  scene.add(mesh);
}}

function viga(x1, x2, y, z) {{
  const largo = Math.abs(x2 - x1);
  const geo = new THREE.BoxGeometry(largo, 0.08, 0.05);
  const mesh = new THREE.Mesh(geo, matViga);
  mesh.position.set((x1+x2)/2, y, z);
  scene.add(mesh);
}}

function cruzZigzag(x, z1, z2, yBase, yTope) {{
  const puntos1 = [new THREE.Vector3(x, yBase, z1), new THREE.Vector3(x, yTope, z2)];
  const puntos2 = [new THREE.Vector3(x, yBase, z2), new THREE.Vector3(x, yTope, z1)];
  [puntos1, puntos2].forEach(p => {{
    const geo = new THREE.BufferGeometry().setFromPoints(p);
    scene.add(new THREE.Line(geo, new THREE.LineBasicMaterial({{color: 0x475569}})));
  }});
}}

for (let corrida = 0; corrida < G.n_corridas; corrida++) {{
  const zBase = corrida * (G.fondo + G.pasillo);
  const zFrente = zBase;
  const zFondo = zBase + G.fondo;

  // Marcos (postes en pares frente/fondo) en cada división de bay
  for (let m = 0; m <= G.n_bays; m++) {{
    const x = m * G.frente;
    poste(x, 0, zFrente, G.altura);
    poste(x, 0, zFondo, G.altura);

    // Cruces (x-bracing) en zigzag entre niveles, en el plano frente-fondo
    for (let n = 0; n < G.niveles.length - 1; n++) {{
      cruzZigzag(x, zFrente, zFondo, G.niveles[n], G.niveles[n+1]);
    }}

    // Cargador extra si frente >= 2804mm: poste intermedio a media altura de cada nivel
    if (G.dos_cargadores && m < G.n_bays) {{
      const xMedio = x + G.frente/2;
      G.niveles.slice(1).forEach(ny => {{
        const geo = new THREE.BoxGeometry(0.05, 0.05, G.fondo);
        const mesh = new THREE.Mesh(geo, matCruz);
        mesh.position.set(xMedio, ny, (zFrente+zFondo)/2);
        scene.add(mesh);
      }});
    }}
  }}

  // Largueros (vigas horizontales) por nivel, uniendo cada par de marcos consecutivos
  G.niveles.slice(1).forEach(ny => {{
    for (let m = 0; m < G.n_bays; m++) {{
      const x1 = m * G.frente, x2 = (m+1) * G.frente;
      viga(x1, x2, ny, zFrente);
      viga(x1, x2, ny, zFondo);
    }}
  }});
}}

function animate() {{
  requestAnimationFrame(animate);
  controls.update();
  renderer.render(scene, camera);
}}
animate();

window.addEventListener('resize', () => {{
  camera.aspect = window.innerWidth/window.innerHeight;
  camera.updateProjectionMatrix();
  renderer.setSize(window.innerWidth, window.innerHeight);
}});
</script>
</body>
</html>"""
