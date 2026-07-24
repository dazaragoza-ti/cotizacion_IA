"""Genera XLSX de despiece a partir del JSON de proyecto.

Uso:  python exportar_xlsx.py datos.json out_dir
Produce:
  - Despiece_<clave>.xlsx     (de datos["materiales"])

La cotización YA NO se genera aquí en XLSX -- ver generators/cotizacion_pdf.py
(reemplazo en PDF, incluye el descuento del Cotizador IA si aplica). La
función cotizacion() se deja abajo por si algo externo todavía la invoca
directamente, pero generar() ya no la llama por default.
"""
from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

AZUL = "1A3A8C"
GRIS = "EFEFEF"
VERDE = "0F7B3F"


def _fecha(datos: dict) -> str:
    return datos.get("fecha") or date.today().strftime("%d/%m/%Y")


def _header(ws, titulo: str, subtitulo: str, ncols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, titulo)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(2, 1, subtitulo).alignment = Alignment(horizontal="center")


def _anchos(ws, anchos: list[int]) -> None:
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fila_encabezado(ws, fila: int, cols: list[str]) -> None:
    for j, h in enumerate(cols, 1):
        c = ws.cell(fila, j, h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=GRIS)


def _descuento_pct(datos: dict) -> float:
    pct = float(datos.get("descuento_pct") or 0)
    if pct > 1:
        pct = pct / 100.0
    return pct


def despiece(datos: dict, out_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Despiece"
    cols = ["Pzas", "Código", "Descripción", "Color", "P. Unit.", "Importe", "Obs."]
    cliente = datos.get("cliente") or "—"
    clave = datos.get("clave") or "PROYECTO"
    _header(
        ws,
        f"DESPIECE — {datos.get('proyecto', '')}",
        f"Clave: {clave}  ·  Cliente: {cliente}  ·  Fecha: {_fecha(datos)}",
        len(cols),
    )
    # Fila de especificación
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(cols))
    ws.cell(3, 1, datos.get("especificacion") or "").font = Font(italic=True, size=9)

    _fila_encabezado(ws, 4, cols)
    r = 5
    subtotal = 0.0
    money = "#,##0.00"
    for m in datos.get("materiales", []) or []:
        pzas = m.get("pzas") or 0
        precio = m.get("precio")
        try:
            pzas_f = float(pzas)
        except (TypeError, ValueError):
            pzas_f = 0.0
        try:
            precio_f = float(precio) if precio is not None else None
        except (TypeError, ValueError):
            precio_f = None
        importe = (pzas_f * precio_f) if precio_f is not None else None
        if importe is not None:
            subtotal += importe

        ws.cell(r, 1, pzas)
        ws.cell(r, 2, m.get("codigo"))
        ws.cell(r, 3, m.get("descripcion"))
        ws.cell(r, 4, m.get("color"))
        if precio_f is not None:
            cu = ws.cell(r, 5, precio_f)
            cu.number_format = money
            ci = ws.cell(r, 6, round(importe or 0, 2))
            ci.number_format = money
        else:
            ws.cell(r, 5, "cotizar")
            ws.cell(r, 6, "—")
        ws.cell(r, 7, m.get("obs", ""))
        r += 1

    # Totales
    r += 1
    ws.cell(r, 5, "Subtotal").font = Font(bold=True)
    cs = ws.cell(r, 6, round(subtotal, 2))
    cs.font = Font(bold=True)
    cs.number_format = money

    pct = _descuento_pct(datos)
    if pct > 0 and subtotal > 0:
        r += 1
        monto = round(subtotal * pct, 2)
        cell_d = ws.cell(r, 5, f"Descuento ({pct * 100:.1f}%)")
        cell_d.font = Font(bold=True, color=VERDE)
        cd = ws.cell(r, 6, -monto)
        cd.font = Font(bold=True, color=VERDE)
        cd.number_format = money
        motivo = datos.get("motivo_descuento") or ""
        if motivo:
            ws.cell(r, 7, motivo)
        r += 1
        ws.cell(r, 5, "TOTAL CON DESCUENTO").font = Font(bold=True, color="FFFFFF")
        ws.cell(r, 5).fill = PatternFill("solid", fgColor=AZUL)
        ct = ws.cell(r, 6, round(subtotal - monto, 2))
        ct.font = Font(bold=True, color="FFFFFF")
        ct.fill = PatternFill("solid", fgColor=AZUL)
        ct.number_format = money
    else:
        r += 1
        ws.cell(r, 5, "TOTAL").font = Font(bold=True, color="FFFFFF")
        ws.cell(r, 5).fill = PatternFill("solid", fgColor=AZUL)
        ct = ws.cell(r, 6, round(subtotal, 2))
        ct.font = Font(bold=True, color="FFFFFF")
        ct.fill = PatternFill("solid", fgColor=AZUL)
        ct.number_format = money

    r += 2
    ws.cell(r, 3, "IVA, flete e instalación se cotizan por separado.").font = Font(italic=True)

    _anchos(ws, [8, 18, 52, 14, 12, 14, 28])
    p = out_dir / f"Despiece_{clave}.xlsx"
    wb.save(p)
    return p


def cotizacion(datos: dict, out_dir: Path) -> Path | None:
    items = datos.get("cotizacion") or []
    if not items:
        # Derivar de los materiales (con o sin precio).
        items = [
            {
                "codigo": m.get("codigo"),
                "descripcion": m.get("descripcion"),
                "cantidad": m.get("pzas"),
                "precio_unitario": m.get("precio"),
            }
            for m in datos.get("materiales", []) or []
        ]
    if not items:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"
    cols = ["Código", "Descripción", "Cant.", "P. Unit.", "Importe"]
    clave = datos.get("clave") or "PROYECTO"
    _header(
        ws,
        f"COTIZACIÓN — {datos.get('proyecto', '')}",
        f"Clave: {clave}  ·  Cliente: {datos.get('cliente') or '—'}  ·  "
        f"Fecha: {_fecha(datos)}  ·  MXN sin IVA (mayoreo)",
        len(cols),
    )
    _fila_encabezado(ws, 4, cols)
    r = 5
    subtotal = 0.0
    for it in items:
        imp = it.get("importe")
        pu = it.get("precio_unitario")
        if imp is None:
            try:
                if pu is not None:
                    imp = float(it.get("cantidad") or 0) * float(pu)
                else:
                    imp = None
            except (TypeError, ValueError):
                imp = None
        if imp is not None:
            subtotal += float(imp)
        ws.cell(r, 1, it.get("codigo"))
        ws.cell(r, 2, it.get("descripcion"))
        ws.cell(r, 3, it.get("cantidad"))
        if pu is not None:
            try:
                cu = ws.cell(r, 4, float(pu))
                cu.number_format = "#,##0.00"
            except (TypeError, ValueError):
                ws.cell(r, 4, "cotizar")
        else:
            ws.cell(r, 4, "cotizar")
        if imp is not None:
            ci = ws.cell(r, 5, float(imp))
            ci.number_format = "#,##0.00"
        else:
            ws.cell(r, 5, "—")
        r += 1

    ws.cell(r, 4, "Subtotal").font = Font(bold=True)
    cs = ws.cell(r, 5, round(subtotal, 2))
    cs.font = Font(bold=True)
    cs.number_format = "#,##0.00"

    pct = _descuento_pct(datos)
    if pct > 0 and subtotal > 0:
        r += 1
        monto = round(subtotal * pct, 2)
        ws.cell(r, 4, f"Descuento ({pct * 100:.1f}%)").font = Font(bold=True, color=VERDE)
        cd = ws.cell(r, 5, -monto)
        cd.font = Font(bold=True, color=VERDE)
        cd.number_format = "#,##0.00"
        r += 1
        ws.cell(r, 4, "TOTAL CON DESCUENTO").font = Font(bold=True)
        ct = ws.cell(r, 5, round(subtotal - monto, 2))
        ct.font = Font(bold=True)
        ct.number_format = "#,##0.00"
    else:
        r += 1
        ws.cell(r, 4, "TOTAL").font = Font(bold=True)
        ct = ws.cell(r, 5, round(subtotal, 2))
        ct.font = Font(bold=True)
        ct.number_format = "#,##0.00"

    ws.cell(r + 2, 2, "IVA, flete e instalación se cotizan por separado.").font = Font(italic=True)
    _anchos(ws, [18, 54, 10, 14, 16])
    p = out_dir / f"Cotizacion_{clave}.xlsx"
    wb.save(p)
    return p


def generar(datos: dict, out_dir) -> list[Path]:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    return [despiece(datos, out_dir)]


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python exportar_xlsx.py datos.json out_dir")
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        datos = json.load(f)
    for p in generar(datos, sys.argv[2]):
        print("XLSX:", p)
